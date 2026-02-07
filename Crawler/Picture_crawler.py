import random
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
import time
from colorama import Fore
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re
import requests
import json
import os
import pandas as pd
import tkinter as tk
from tkinter import messagebox, scrolledtext, ttk
import threading
import queue


class CarConfigApp:
    def __init__(self, root):
        self.root = root
        self.root.title("汽车配置信息下载工具")
        self.root.geometry("800x700")
        # self.root.iconbitmap('15boe-37g29-001.ico')
        self.root.config(bg="#F8F8FF")

        # 界面元素
        self.brand_label = tk.Label(root, text="请输入汽车品牌:")
        self.brand_label.pack(pady=10)

        self.brand_entry = tk.Entry(root, width=50)
        self.brand_entry.pack(pady=5)

        self.search_btn = tk.Button(root, text="查询品牌", command=self.start_brand_search)
        self.search_btn.pack(pady=5)

        self.log_text = scrolledtext.ScrolledText(root, width=100, height=15)
        self.log_text.pack(pady=10)

        self.series_listbox = tk.Listbox(root, width=80, height=10)
        self.series_listbox.pack(pady=10)

        self.progress_bar = ttk.Progressbar(
            root,
            length=400,
            mode='determinate',
            maximum=100
        )
        self.progress_bar.pack(pady=10, padx=20)

        self.download_btn = tk.Button(root, text="下载选中车型配置", command=self.start_download)
        self.download_btn.pack(pady=20)
        self.download_btn.config(state=tk.DISABLED)

        self.series_dict = {}
        self.current_brand = None
        self.log_queue = queue.Queue()
        self.update_log()

    def start_brand_search(self):
        brand = self.brand_entry.get().strip()
        if not brand:
            messagebox.showwarning("输入错误", "请输入汽车品牌")
            return

        self.log_text.delete(1.0, tk.END)
        self.series_listbox.delete(0, tk.END)
        self.download_btn.config(state=tk.DISABLED)

        threading.Thread(target=self.get_brand_info, args=(brand,)).start()

    def get_brand_info(self, brand):
        try:
            response = get_band_response()
            band_pattern = f"<a href=([^>]*?)><i[^>]*?></i>{brand}<em>"
            band_info = re.search(band_pattern, response.text)

            if not band_info:
                self.log_queue.put(f"品牌 {brand} 不存在，请重新输入")
                return

            band_href = band_info.group(1)
            band_id = re.findall(r'/price/brand-(\d+).html', band_href)[0]
            self.current_brand = brand
            self.log_queue.put(f"{brand} 品牌id为：{band_id}")

            resp_brand = get_band_response(brand_id=band_id)
            self.parse_series_gui(resp_brand)
        except Exception as e:
            self.log_queue.put(f"错误：{str(e)}")

    def parse_series_gui(self, response):
        html = re.findall(r'document.writeln\("(.*)"\)', response.text)
        html = "".join(html)
        soup = BeautifulSoup(html, "html.parser")
        data_list = soup.select(".current > dl > dd > a")
        still_sell = [i for i in data_list if "停售" not in i.get_text(strip=True)]

        self.log_queue.put(f"共找到{len(data_list)}个车型，其中在售{len(still_sell)}个")

        self.series_dict = {}
        for idx, data in enumerate(still_sell, 1):
            series_name = data.contents[0].text.strip()
            href = data.get("href")
            series_id = re.findall(r'/price/series-(\d+).html', href)[0]
            self.series_dict[series_id] = series_name
            self.log_queue.put(f"序号：{idx}\t车型：{series_name}\tID：{series_id}")
            self.series_listbox.insert(tk.END, f"{series_name} (ID: {series_id})")

        if self.series_dict:
            self.download_btn.config(state=tk.NORMAL)
        else:
            self.log_queue.put("该品牌下没有在售车型")

    def start_download(self):
        selected = self.series_listbox.curselection()
        if not selected:
            messagebox.showwarning("选择错误", "请选择一个车型")
            return

        series_id = list(self.series_dict.keys())[selected[0]]
        series_name = self.series_dict[series_id]

        threading.Thread(target=self.download_series, args=(series_id, series_name)).start()

    def download_series(self, series_id, series_name):
        try:
            self.log_queue.put(f"---正在下载{self.current_brand}-{series_name}，ID：{series_id}")
            response = get_response(series_id)

            if "抱歉" in response.text and "暂无相关数据" in response.text:
                self.log_queue.put("该系列车暂无配置信息")
                return

            config_dic = json.loads(response.text)
            all_info = get_car_config(config_dic)
            excel_name = f"{self.current_brand}_{series_name}.xlsx"

            save_to_excel(all_info, self.current_brand, excel_name)
            self.log_queue.put(f"配置下载完成，保存到：{self.current_brand}/{excel_name}")

            self.log_queue.put(f"开始下载{series_name}的图片...")
            self.download_images(series_id, series_name)
        except Exception as e:
            self.log_queue.put(f"下载失败：{str(e)}")

    def download_images(self, series_id, series_name):
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
                "Referer": "https://www.autohome.com.cn/cars/"
            }
            base_url = f"https://car.autohome.com.cn/pic/series/{series_id}-1-p"
            page = 1
            max_pages = 20  # 最大请求页数限制
            all_image_pages = []  # 存储最终唯一链接的列表
            unique_pages = set()  # 用于去重

            while page <= max_pages:
                url = f"{base_url}{page}.html"
                response = requests.get(url, headers=headers, timeout=10)
                if response.status_code != 200:
                    self.log_queue.put(f"请求失败：{url}，状态码：{response.status_code}")
                    break
                soup = BeautifulSoup(response.text, 'html.parser')
                pattern = re.compile(r'/cars/imgs-.*?\.html')
                links = soup.find_all('a', href=pattern)
                found_links = len(links)
                self.log_queue.put(f"解析第{page}页图片列表，共找到{found_links}个图片页面")
                if found_links == 0:
                    self.log_queue.put(f"第{page}页无有效图片链接，停止请求")
                    break
                for link in links:
                    href = link['href']
                    if href.startswith('http'):
                        img_page_url = href
                    else:
                        img_page_url = "https://car.autohome.com.cn" + href
                    # 检查是否已存在
                    if img_page_url not in unique_pages:
                        unique_pages.add(img_page_url)
                        all_image_pages.append(img_page_url)
                page += 1
                time.sleep(random.uniform(1, 2))

            # 解析图片URL
            all_image_urls = []
            for img_page_url in all_image_pages:
                response = requests.get(img_page_url, headers=headers, timeout=10)
                if response.status_code != 200:
                    continue
                soup = BeautifulSoup(response.text, 'html.parser')
                img_pattern = re.compile(r'https://car\d+\.autoimg\.cn/cardfs/product/.*?\.(jpg|png|gif)')
                img_tags = soup.find_all('img', src=img_pattern)
                for img_tag in img_tags:
                    img_url = img_tag['src']
                    all_image_urls.append(img_url)
                self.log_queue.put(f"解析页面 {img_page_url}，找到{len(img_tags)}张图片")

            # 创建文件夹
            folder_path = os.path.join(self.current_brand, series_name)
            os.makedirs(folder_path, exist_ok=True)
            self.log_queue.put(f"文件夹路径：{folder_path}")

            total_images = len(all_image_urls)
            if total_images == 0:
                self.log_queue.put("未找到可下载的图片")
                self.finish_download()
                return

            self.set_progress_max(total_images)
            self.disable_ui()
            completed = 0

            for idx, img_url in enumerate(all_image_urls):
                filename = os.path.basename(img_url)
                file_path = os.path.join(folder_path, filename)
                if os.path.exists(file_path):
                    self.log_queue.put(f"跳过已存在的图片：{filename}")
                    completed += 1
                    self.update_progress(completed)
                    continue

                success = self.download_single_image(img_url, series_id, series_name)
                if success:
                    completed += 1
                    self.log_queue.put(f"下载成功：{filename}")
                else:
                    self.log_queue.put(f"下载失败：{filename}")

                self.update_progress(completed)
                time.sleep(random.uniform(0.5, 1.5))

            self.log_queue.put(f"所有图片下载完成，保存到：{folder_path}")
            self.finish_download()

        except Exception as e:
            self.log_queue.put(f"图片下载错误：{str(e)}")
            self.finish_download()


    def download_single_image(self, img_url, series_id, series_name):
        try:
            headers = {"User-Agent": UserAgent().random}
            response = requests.get(img_url, headers=headers, stream=True, timeout=10)
            if response.status_code == 200:
                folder_path = os.path.join(self.current_brand, series_name)
                os.makedirs(folder_path, exist_ok=True)
                filename = os.path.basename(img_url)
                file_path = os.path.join(folder_path, filename)
                with open(file_path, 'wb') as f:
                    for chunk in response.iter_content(1024):
                        f.write(chunk)
                return True
            else:
                return False
        except Exception as e:
            self.log_queue.put(f"下载图片 {img_url} 失败：{str(e)}")
            return False

    def set_progress_max(self, max_val):
        self.progress_bar['maximum'] = max_val
        self.progress_bar['value'] = 0

    def update_progress(self, completed):
        self.progress_bar['value'] = completed
        self.root.update_idletasks()

    def disable_ui(self):
        self.series_listbox.config(state=tk.DISABLED)
        self.download_btn.config(state=tk.DISABLED)

    def finish_download(self):
        self.progress_bar['value'] = 0
        self.series_listbox.config(state=tk.NORMAL)
        self.download_btn.config(state=tk.NORMAL)

    def update_log(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                self.log_text.insert(tk.END, msg + "\n")
                self.log_text.see(tk.END)
        except queue.Empty:
            pass
        self.root.after(100, self.update_log)

def get_band_response(brand_id="0"):
    num = 1
    while True:
        headers = {"user-agent": UserAgent().random}
        url = "https://car.autohome.com.cn/AsLeftMenu/As_LeftListNew.ashx"
        params = {"typeId": "1", "brandId": brand_id, "fctId": "0", "seriesId": "0"}
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            return response
        else:
            if num >= 5:
                raise Exception("请求失败，退出程序")
            time.sleep(1)
            num += 1

def get_response(series_id="0"):
    num = 1
    while True:
        headers = {"user-agent": UserAgent().random}
        url = "https://car-web-api.autohome.com.cn/car/param/getParamConf"
        params = {"mode": "1", "site": "1", "seriesid": series_id}
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            return response
        else:
            if num >= 5:
                raise Exception("请求失败，退出程序")
            time.sleep(1)
            num += 1

def get_car_config(config_dic):
    allconfig = []
    configname_list = []

    for title in config_dic['result']['titlelist']:
        for item in title['items']:
            configname_list.append(item['itemname'])
    allconfig.append(configname_list)

    for data in config_dic['result']['datalist']:
        configvalue_list = []
        for valueitem in data['paramconflist']:
            if valueitem.get('itemname'):
                configvalue_list.append(valueitem['itemname'])
            elif not valueitem.get('sublist'):
                configvalue_list.append('-')
            else:
                values = [f"{v['value']}{v['name']}" for v in valueitem['sublist']]
                configvalue_list.append('\n'.join(values))
        allconfig.append(configvalue_list)

    return allconfig

def save_to_excel(data, folder, filename):
    if not os.path.exists(folder):
        os.mkdir(folder)
    df = pd.DataFrame(data)
    excel_path = os.path.join(folder, filename)
    df.T.to_excel(excel_path, index=False, header=False)

    wb = load_workbook(excel_path)
    sheet = wb.active
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    for col in range(1, df.shape[0] + 1):
        sheet.column_dimensions[chr(64 + col)].width = 20
    wb.save(excel_path)

if __name__ == '__main__':
    root = tk.Tk()
    app = CarConfigApp(root)
    root.mainloop()